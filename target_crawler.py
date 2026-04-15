import requests
import json
import time
import urllib.parse
import re
import html
import os
import random
import math
import pandas as pd
from datetime import datetime, timedelta
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
import urllib3

# 尝试导入进度条
try:
    from tqdm import tqdm
except ImportError:
    def tqdm(iterator, **kwargs): return iterator

# 禁用SSL警告
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# ==================== ✅ 随机访客ID生成器 (代码的"无痕模式") ====================
def generate_visitor_id():
    """生成一个随机的32位大写十六进制字符串，模拟Target的新访客，防止被Cookie黑名单拦截"""
    return ''.join(random.choices('0123456789ABCDEF', k=32))

# ==================== 全局配置 ====================
# 每次运行脚本，自动获取一个干净的全新身份
VISITOR_ID = generate_visitor_id()

# API URL (已使用 f-string 动态注入生成的 VISITOR_ID)
BASE_API_URL = f"https://redsky.target.com/redsky_aggregations/v1/web/plp_search_v2?category=5xtd3&count=24&default_purchasability_filter=true&include_sponsored=true&include_review_summarization=true&offset=0&page=%2Fc%2F5xtd3&platform=desktop&pricing_store_id=1121&spellcheck=true&store_ids=1121%2C267%2C311%2C1098%2C1502&visitor_id={VISITOR_ID}&scheduled_delivery_store_id=1121&zip=95628&key=9f36aeafbe60771e321a7cc95a78140772ab3e96&channel=WEB&include_dmc_dmr=true&useragent=Mozilla%2F5.0+%28Windows+NT+10.0%3B+Win64%3B+x64%29+AppleWebKit%2F537.36+%28KHTML%2C+like+Gecko%29+Chrome%2F143.0.0.0+Safari%2F537.36+Edg%2F143.0.0.0"

API_KEY = "9f36aeafbe60771e321a7cc95a78140772ab3e96"
REVIEW_API_KEY = "c6b68aaef0eac4df4931aae70500b7056531cb37" 
TARGET_STORE_ID = "1121"
TARGET_ZIP_CODE = "95628"
REQUESTS_TIMEOUT = 15

MAX_WORKERS_LISTING = 28
MAX_WORKERS_DETAIL = 28 
MAX_API_OFFSET_LIMIT = 1199
MAX_REVIEW_PAGES = 5  # 限制每个商品的评论抓取页数

# 全局 Headers
HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Accept": "application/json, text/plain, */*",
    "Referer": "https://www.target.com/",
    "Accept-Language": "en-US,en;q=0.9", # 增加一层伪装
}

# 目标品类列表
TARGET_CATEGORIES = {
    "Adhesive Undergarment Covers", "Apparel Accessories Sets", "Athletic Bodysuits", 
    "Athletic Bottoms", "Athletic Dresses", "Athletic Rompers and Jumpsuits", 
    "Athletic Tops", "Athletic Wear Sets", "Beauty Tools and Sets", "Bra Accessories", 
    "Bras", "Coats and Jackets", "Combination Bottoms", "Compression Hosiery", 
    "Coordinate Sets", "Costume Full Body Apparel", "Dresses", "Hosiery", 
    "Legwear Collections", "Leotards and Bodysuits", "License Graphic Shirts", 
    "Maternity Support Garments", "One Piece Pajamas", "One Piece Swimsuits", 
    "Outerwear Bottoms", "Pajama Bottoms", "Pajama Sets", "Pajama Tops", "Pants", 
    "Rompers and Jumpsuits", "Shapewear", "Shirts", "Shorts", "Skirts", 
    "Slips and Lingerie", "Socks", "Sweaters and Shawls", "Sweatshirts", 
    "Swimsuit Bottoms", "Swimsuit Cover Ups", "Swimsuit Tops", "Swimwear Sets", 
    "Top Sets", "Undergarment Sets", "Underwear", "Vests"
}

# 超大品牌名单
MEGA_BRANDS = {"Disney", "Instant Message", "INSPIRE CHIC", "LOST GODS", "Marvel", "MeMoi"}

# 配置 Session
session = requests.Session()
retry = Retry(total=3, backoff_factor=0.5, status_forcelist=[400, 403, 429, 500, 502, 503, 504])
adapter = HTTPAdapter(max_retries=retry)
session.mount("https://", adapter)
session.headers.update(HEADERS)

# ==================== PART 1: 线上列表爬取类 ====================

class TargetCrawler:
    def __init__(self, base_url: str):
        self.base_url = base_url
        self.parsed_url = urllib.parse.urlparse(base_url)
        self.base_query_params = urllib.parse.parse_qs(self.parsed_url.query)

    def fetch_json(self, url: str):
        try:
            resp = session.get(url, timeout=15, verify=False)
            if resp.status_code == 400: return {"_error": 400}
            resp.raise_for_status()
            return resp.json()
        except: return {}

    def extract_cat_id(self, path: str):
        match = re.search(r'N-([a-zA-Z0-9]+)', path)
        return match.group(1) if match else None

    def extract_product_data(self, products):
        extracted = []
        for p in products:
            tcin = p.get("tcin") or p.get("parent", {}).get("tcin")
            if tcin:
                if 'tcin' not in p: p['tcin'] = tcin
                extracted.append(p)
        return extracted

    def _build_url(self, params):
        return urllib.parse.urlunparse((
            self.parsed_url.scheme, self.parsed_url.netloc, self.parsed_url.path,
            self.parsed_url.params, urllib.parse.urlencode(params, doseq=True), self.parsed_url.fragment
        ))

    def get_target_category_ids(self):
        print("🔍 [阶段1] 请求 Base URL，解析 Item Type 列表...")
        data = self.fetch_json(self.base_url)
        matched = []
        resp = data.get("data", {}).get("search", {}).get("search_response", {}) or \
               data.get("data", {}).get("search_response", {})
        
        target_facet = next((f for f in resp.get("facet_list", []) if f.get("facet_id") == "d_item_type_apparel_accessories"), None)
        if not target_facet:
            print("⚠️ 警告：未找到细分分类 Facet。")
            print("🔄 启动兜底策略：直接使用 [Women's Clothing (5xtd3)] 主分类进行全量扫描...")
            return [{"name": "Women's Clothing", "id": "5xtd3"}]

        for option in target_facet.get("options", []):
            name = option.get("display_name", "").strip()
            if name in TARGET_CATEGORIES:
                cat_id = self.extract_cat_id(option.get("url", "")) or option.get("value")
                if cat_id: matched.append({"name": name, "id": cat_id})
        
        print(f"✅ 匹配到 {len(matched)} 个目标分类 ID。")
        return matched

    def get_brands_for_category(self, cat_id, cat_name):
        params = self.base_query_params.copy()
        params.update({'page': [f"/c/{cat_id}"], 'category': [cat_id], 'offset': ['0']})
        url = self._build_url(params)
        data = self.fetch_json(url)
        brand_list = []
        
        resp = data.get("data", {}).get("search", {}).get("search_response", {}) or \
               data.get("data", {}).get("search_response", {})
        
        for facet in resp.get("facet_list", []):
            if facet.get("facet_id") == "d_brand_all":
                for option in facet.get("options", []):
                    name = option.get("display_name", "").strip()
                    path = option.get("url", "").strip()
                    if name:
                        brand_list.append({"brand_name": name, "brand_path": path, "cat_name": cat_name, "cat_id": cat_id})
        return brand_list

    def _crawl_loop(self, cat_id, sort_by=None):
        collected_products = []
        offset = 0
        while True:
            if offset > MAX_API_OFFSET_LIMIT: break
            params = self.base_query_params.copy()
            params.update({'page': [f"/c/{cat_id}"], 'category': [cat_id], 'offset': [str(offset)]})
            if sort_by: params.update({'sort_by': [sort_by]})

            url = self._build_url(params)
            data = self.fetch_json(url)
            if not data or data.get("_error") == 400: break

            prods = data.get("data", {}).get("search", {}).get("products", []) or \
                    data.get("data", {}).get("data", {}).get("product_summaries", [])
            if not prods: break
            
            collected_products.extend(self.extract_product_data(prods))
            if len(prods) < 24: break
            offset += 24
        return collected_products

    def crawl_task_logic(self, task_info):
        brand_name = task_info['brand_name']
        brand_cat_id = self.extract_cat_id(task_info['brand_path'])
        if not brand_cat_id: return []

        products = []
        if brand_name in MEGA_BRANDS:
            for sort_val in ["newest", "PriceLow", "PriceHigh", "RatingHigh", "BestSeller"]:
                products.extend(self._crawl_loop(brand_cat_id, sort_by=sort_val))
        else:
            products = self._crawl_loop(brand_cat_id)
            if len(products) >= 1100:
                 products.extend(self._crawl_loop(brand_cat_id, sort_by="newest"))
        
        seen = set()
        unique_products = []
        for p in products:
            t = p.get('tcin')
            if t and t not in seen:
                seen.add(t)
                unique_products.append(p)
                
        return unique_products

# ==================== PART 2: 核心解析逻辑 ====================

def clean_text(text):
    if text is None: return None
    text = html.unescape(str(text))
    text = re.sub(r'<[^>]+>', '', text)
    return text.strip()

def format_bought_count(raw_text):
    if not raw_text: return None
    try:
        match = re.search(r'([\d\.]+)\s*([kKmM]?)\s*\+?', str(raw_text))
        if match:
            num_str, unit = match.groups()
            number = float(num_str)
            if unit:
                if unit.lower() == 'k': number *= 1000
                elif unit.lower() == 'm': number *= 1_000_000
            return int(number)
    except: pass
    return None

# ==================== ✅ 已替换：第二份代码 购买次数逻辑（累加所有子商品） ====================
def extract_sales_velocity_split(product):
    total_sales_num = 0
    trend_tags = []
    all_trends = set()

    try:
        # 遍历所有子商品
        children = product.get("children", [])
        if not children:
            children = [product]

        for child in children:
            item_node = child.get('item', {})
            enrichment_node = item_node.get('enrichment', {})
            search_targets = [child, item_node, enrichment_node]
            child_sales = None

            # 提取单个子商品的购买次数
            for target in search_targets:
                cues = target.get("desirability_cues", [])
                if not cues: continue
                for cue in cues:
                    if cue.get("code") == "social_proofing":
                        raw_msg = cue.get("display", "") or cue.get("text", "")
                        if raw_msg:
                            child_sales = format_bought_count(raw_msg)
                            break
                if child_sales is not None:
                    break

            # 兜底字段
            if child_sales is None:
                raw_msg_old = enrichment_node.get("sales_velocity_message", "").strip()
                if raw_msg_old:
                    child_sales = format_bought_count(raw_msg_old)

            # 累加
            if child_sales is not None and child_sales > 0:
                total_sales_num += child_sales

            # 收集热度标签（去重）
            for target in search_targets:
                orns = target.get("ornaments", [])
                for orn in orns:
                    display = str(orn.get('display', ''))
                    if display and 'bought' not in display.lower():
                        all_trends.add(display)

        trend_str = " | ".join(sorted(all_trends)) if all_trends else ""
        return total_sales_num if total_sales_num > 0 else None, trend_str
    except:
        return None, ""

def extract_ratings(product):
    avg, count, sec_ratings = 0.0, 0, ""
    try:
        stats = product.get('ratings_and_reviews', {}).get('statistics', {}).get('rating', {})
        if not stats: 
            stats = product.get('parent', {}).get('ratings_and_reviews', {}).get('statistics', {}).get('rating', {})
        avg = float(stats.get('average', 0.0))
        count = int(stats.get('count', 0))
        sec_avgs = stats.get('secondary_averages', [])
        if sec_avgs and isinstance(sec_avgs, list):
            parts = [f"{i.get('label') or i.get('id')}: {i.get('value')}" for i in sec_avgs if i.get('value') is not None]
            sec_ratings = " | ".join(parts)
    except: pass
    return avg, count, sec_ratings

def extract_detailed_features_dict(product):
    features = {}
    try:
        raw_bullets = product.get('item', {}).get('product_description', {}).get('bullet_descriptions', [])
        for b in raw_bullets:
            if not b: continue
            clean_b = re.sub(r'<[^>]+>', '', str(b)).strip()
            clean_b = html.unescape(clean_b)
            if ':' in clean_b:
                parts = clean_b.split(':', 1)
                key = parts[0].strip()
                val = parts[1].strip()
                if key and val: features[key] = val
    except: pass
    return features

def extract_material_summary(product):
    material = ''
    try:
        item = product.get('item', {})
        p_desc = item.get('product_description', {}) if item else {}
        for bullet in p_desc.get('bullet_descriptions', []):
            clean_b = re.sub(r'<[^>]+>', '', str(bullet)).strip()
            if 'Material:' in clean_b:
                material = clean_b.split(':', 1)[1].strip(); break
        if not material:
            for attr in item.get('wellness_merchandise_attributes', []):
                val = str(attr.get('value_name', ''))
                if 'Recycled' in val: material = val; break
    except: pass
    return material[:100]

# ==================== ✅ 修复版：颜色提取（完美支持你给的JSON）+ 新增尺码提取 ====================
def get_variation_from_json(product_data):
    """最稳健的方法：直接读取官方定义的 Themes，同时提取颜色+尺码"""
    c_set = set()
    s_set = set()
    sources = [product_data, product_data.get("parent", {})]
    for data in sources:
        if not data: continue
        var_summary = data.get("variation_summary", {})
        themes = var_summary.get("themes", [])
        for theme in themes:
            t_name = theme.get("name", "").lower()
            swatches = theme.get("swatches", [])
            values = [clean_text(s.get("value")) for s in swatches if s.get("value")]
            if "color" in t_name or "pattern" in t_name:
                c_set.update(values)
            # 新增：提取尺码
            elif "size" in t_name:
                s_set.update(values)
    return list(c_set), list(s_set)

def extract_color_from_title_fallback(title):
    """
    🔥 终极修复版：完美识别
    1. Brown/Pink/Ivory → 自动保留组合色，不拆分
    2. Light Blue Striped → 完整识别
    3. 所有带条纹/图案/组合的颜色都能正确提取
    """
    if not title:
        return []
    
    title = html.unescape(str(title)).strip()
    
    # 匹配 Target 服饰标题格式：品牌™ 后 → 颜色 → 尺码
    pattern = r"""
        .*?
        -\s*
        .*?™?\s*
        ([A-Za-z\s\/\-]+?(?:Striped|Heather|Print|Pattern|Solid)?)
        \s*
        (?:XS|S|M|L|XL|XXL|XXXL|Plus|Size|\d+-\d+|\d+)?
        \s*$
    """

    match = re.search(pattern, title, re.IGNORECASE | re.VERBOSE)
    if not match:
        return []

    color_str = match.group(1).strip()
    if not (2 <= len(color_str) <= 60):
        return []

    # 直接返回完整颜色，不拆分 /，完美适配你的JSON
    return [color_str]

# ==================== ✅ 新增：尺码兜底提取函数 ====================
def extract_size_fallback(product):
    """从商品属性/描述兜底提取尺码"""
    size_set = set()
    try:
        item = product.get("item", {})
        # 从商品属性提取
        attrs = item.get("product_characteristics", {}).get("attributes", [])
        for attr in attrs:
            if attr.get("name", "").lower() in ["size", "sizing"]:
                val = attr.get("value", "")
                if val:
                    size_set.add(val.strip())
        
        # 从描述提取
        desc = item.get("product_description", {}).get("title", "")
        if desc:
            # 匹配常见尺码
            size_pattern = re.compile(r'\b(XS|S|M|L|XL|XXL|XXXL|OS|One Size|\d+-\d+|\d+W?)\b', re.I)
            matches = size_pattern.findall(desc)
            for m in matches:
                size_set.add(m.strip().upper())
    except:
        pass
    return list(size_set)

def crawl_delivery_date(tcin):
    fulfill_url = "https://redsky.target.com/redsky_aggregations/v1/web/product_fulfillment_v1"
    date_result = "无物流信息"
    try:
        params = {
            "key": API_KEY,
            "channel": "WEB",
            "tcin": tcin,
            "zip": TARGET_ZIP_CODE,
            "store_id": TARGET_STORE_ID,
            "required_store_id": TARGET_STORE_ID,
            "scheduled_delivery_store_id": TARGET_STORE_ID
        }
        headers = HEADERS.copy()
        headers['Referer'] = f"https://www.target.com/p/-/A-{tcin}"
        
        resp = session.get(fulfill_url, headers=headers, params=params, timeout=REQUESTS_TIMEOUT)
        if resp.status_code in [200, 206]:
            data = resp.json()
            root = data.get('data', {}).get('product', {}) or data.get('data', {})
            services = root.get('fulfillment', {}).get('shipping_options', {}).get('services', [])
            if services and len(services) > 0:
                display_text = services[0].get('display_text', '')
                min_date = services[0].get('min_delivery_date', '')
                if display_text:
                    date_result = display_text
                elif min_date:
                    date_result = f"{min_date}"
    except Exception:
        pass
    return date_result

# 增强版评论抓取逻辑
def crawl_reviews_enhanced(tcin, product_title, product_buy_url):
    all_reviews = []
    page = 0
    has_more = True
    headers = HEADERS.copy()
    headers["Referer"] = f"https://www.target.com/p/A-{tcin}"

    while has_more:
        try:
            time.sleep(random.uniform(0.3, 0.8))
            review_request_url = (
                f"https://r2d2.target.com/ratings_reviews_api/v1/summary"
                f"?key={REVIEW_API_KEY}&hasOnlyPhotos=false"
                f"&includes=reviews%2CreviewsWithPhotos%2Centities%2Cmetadata%2Cstatistics"
                f"&page={page}&reviewedId={tcin}&reviewType=PRODUCT&size=8&sortBy=most_recent&verifiedOnly=false"
            )
            res = session.get(review_request_url, headers=headers, timeout=10, verify=False)
            if res.status_code != 200: break
            review_data = res.json()
            results = review_data.get("reviews", {}).get("results", [])
            if not results: break

            for review in results:
                all_reviews.append({
                    "商品TCIN": tcin,
                    "商品标题": product_title,
                    "商品购买链接": product_buy_url,
                    "评论标题": review.get("title", "无标题"),
                    "评论内容": review.get("text", "").replace("\n", " "),
                    "星级": review.get("Rating", ""),
                    "作者": review.get("author", {}).get("nickname", "匿名用户"),
                    "是否验证购买": "Yes" if review.get("isVerified") else "No",
                    "有用数": review.get("votes_up", 0),
                    "提交时间": review.get("submitted_at", ""),
                    "评论带图": "Yes" if len(review.get("photos", [])) > 0 else "No"
                })
            page += 1
            total_results = review_data.get("reviews", {}).get("total_results", 0)
            if len(all_reviews) >= total_results or page >= MAX_REVIEW_PAGES:
                has_more = False
        except: has_more = False
    return all_reviews

# 解析价格区间辅助函数
def parse_price_range_values(price_str):
    """
    解析价格字符串，返回 (min, max, is_range)
    "$10" -> (10.0, 10.0, False)
    "$10 - $20" -> (10.0, 20.0, True)
    """
    if not price_str: return 0.0, 0.0, False
    
    clean_str = str(price_str).replace('$', '').replace(',', '').strip()
    
    if '-' in clean_str:
        try:
            parts = clean_str.split('-')
            min_val = float(parts[0].strip())
            max_val = float(parts[1].strip())
            return min_val, max_val, True
        except: return 0.0, 0.0, False
    else:
        try:
            val = float(clean_str)
            return val, val, False
        except: return 0.0, 0.0, False

def process_single_product(task_data):
    tcin = task_data['tcin']
    plp_data = task_data.get('plp_data', {}) 
    
    detail_url = (
        f"https://redsky.target.com/redsky_aggregations/v1/web/pdp_client_v1"
        f"?key={API_KEY}&tcin={tcin}&is_bot=false"
        f"&store_id={TARGET_STORE_ID}"
        f"&pricing_store_id={TARGET_STORE_ID}&has_pricing_store_id=true"
        f"&zip={TARGET_ZIP_CODE}"
        f"&has_financing_options=true&include_obsolete=true"
        f"&visitor_id={VISITOR_ID}&skip_personalized=true"
        f"&skip_variation_hierarchy=true&channel=WEB&page=%2Fp%2FA-{tcin}"
    )
    
    try:
        resp = session.get(detail_url, timeout=20, verify=False)
        pdp_data = {}
        if resp.status_code == 200:
            json_data = resp.json()
            pdp_data = json_data.get("data", {}).get("product", {})
        
        if not pdp_data and plp_data: pdp_data = plp_data
        if not pdp_data: return None, []
        
        pdp_item = pdp_data.get("item", {})
        p_desc = pdp_item.get("product_description", {})
        p_enrich = pdp_item.get("enrichment", {})
        p_brand = pdp_item.get("primary_brand", {})
        
        title = clean_text(p_desc.get("title")) or clean_text(pdp_data.get("title")) or \
                clean_text(plp_data.get("item", {}).get("product_description", {}).get("title"))
        
        clean_bullets = [clean_text(b) for b in p_desc.get("soft_bullets", {}).get("bullets", []) if b]
        feature_dict = extract_detailed_features_dict(pdp_data) 

        # ==================== 1. 价格 & 折扣 & 清仓状态 ====================
        price_node = pdp_data.get('price', {}) or plp_data.get('price', {})

        formatted_current = clean_text(price_node.get('formatted_current_price'))
        formatted_comparison = clean_text(price_node.get('formatted_comparison_price'))

        # ✅ 新增：清仓状态识别
        price_type = str(price_node.get('formatted_current_price_type', '')).lower()
        is_clearance = "Yes" if "clearance" in price_type else "No"

        curr_min, curr_max, curr_is_range = parse_price_range_values(formatted_current)
        if not formatted_comparison and price_node.get('reg_retail'):
            reg_val = float(price_node.get('reg_retail'))
            reg_min, reg_max, reg_is_range = reg_val, reg_val, False
        else:
            reg_min, reg_max, reg_is_range = parse_price_range_values(formatted_comparison)

        out_current_price = formatted_current  
        out_retail = None  
        out_reg = None     
        out_save_amt = None
        out_discount_pct = None
        out_max_discount = None
        is_promotion = 'No'
        
        if curr_is_range:
            out_retail = formatted_current
        elif curr_max > 0:
            out_retail = curr_max 
            if not out_current_price: out_current_price = f"${out_retail}"

        if reg_max > 0:
            if reg_is_range:
                out_reg = formatted_comparison
            else:
                out_reg = reg_max

        has_discount = False
        if reg_max > 0 and curr_max > 0:
            if reg_min > curr_min or reg_max > curr_max:
                has_discount = True
                is_promotion = 'Yes'

        if has_discount:
            save_low = max(0, reg_min - curr_min)
            pct_low = save_low / reg_min if reg_min > 0 else 0
            save_high = max(0, reg_max - curr_max)
            pct_high = save_high / reg_max if reg_max > 0 else 0
            
            out_max_discount = max(pct_low, pct_high)
            
            if curr_is_range or reg_is_range:
                if save_low == save_high:
                    out_save_amt = f"${save_low:.2f}"
                else:
                    out_save_amt = f"${min(save_low, save_high):.2f} - ${max(save_low, save_high):.2f}"
                
                p_low_str = f"{int(pct_low * 100)}%"
                p_high_str = f"{int(pct_high * 100)}%"
                
                if abs(pct_low - pct_high) < 0.01: 
                    out_discount_pct = p_low_str
                else:
                    p_min = min(pct_low, pct_high)
                    p_max = max(pct_low, pct_high)
                    out_discount_pct = f"{int(p_min*100)}% - {int(p_max*100)}%"
            else:
                out_save_amt = save_low 
                out_discount_pct = pct_low 
                out_max_discount = pct_low

        else:
            out_reg = None 
            out_save_amt = None
            out_discount_pct = None
            out_max_discount = None

        if not has_discount and price_node.get('save_dollar'):
            try:
                s_val = float(price_node.get('save_dollar'))
                if s_val > 0:
                    is_promotion = 'Yes'
                    out_save_amt = s_val
                    if out_retail and isinstance(out_retail, float):
                        out_reg = out_retail + s_val
                        out_discount_pct = s_val / out_retail
                        out_max_discount = out_discount_pct
            except: pass

        promotions = pdp_data.get('promotions', []) or plp_data.get('promotions', [])
        promo_labels = []
        for promo in promotions:
            if isinstance(promo, dict):
                msg = promo.get('call_out_message')
                if msg:
                    promo_labels.append(msg)
                    is_promotion = 'Yes'
        if promo_labels:
            clean_bullets.insert(0, f"[促销] {' | '.join(promo_labels)}")

        # 2. 颜色+尺码提取（✅ 已新增尺码）
        color_list, size_list = get_variation_from_json(pdp_data)
        if not color_list and plp_data:
             c_plc, s_plc = get_variation_from_json(plp_data)
             color_list.extend(c_plc)
             size_list.extend(s_plc)

        # 标题兜底颜色
        if not color_list and title:
            title_colors = extract_color_from_title_fallback(title)
            color_list.extend(title_colors)
        
        # 兜底尺码
        if not size_list:
            size_list = extract_size_fallback(pdp_data)
        
        color_list = sorted(list(set(color_list)))
        size_list = sorted(list(set(size_list)))
        
        color_result = f"[共{len(color_list)}种] {', '.join(color_list)}" if color_list else ""
        # ✅ 新增：尺码汇总格式化
        size_result = f"[共{len(size_list)}种] {', '.join(size_list)}" if size_list else "无尺码"

        # 3. 其他信息
        delivery_date = crawl_delivery_date(tcin)
        rating_avg, rating_count, sec_ratings = extract_ratings(pdp_data)
        material_summary = extract_material_summary(pdp_data)
        sales_count_num, _ = extract_sales_velocity_split(pdp_data)
        
        # ==================== ✅ 修正版：精准新品判断 ====================
        is_new = 'No'

        ribbons = pdp_item.get('ribbons', [])
        for ribbon in ribbons:
            if isinstance(ribbon, str):
                r_upper = ribbon.upper()
                if 'NEW' in r_upper:
                    if any(key in r_upper for key in ['COLOR', 'SIZE', 'FABRIC', 'PATTERN']):
                        continue
                    is_new = 'Yes'
                    break

        if is_new == 'No':
            for node in [pdp_data, pdp_item, p_enrich]:
                cues = node.get("desirability_cues", [])
                for cue in cues:
                    code = cue.get("code", "")
                    if code in ["new_item", "new_arrival", "new_product"]:
                        is_new = 'Yes'
                        break
                if is_new == 'Yes':
                    break

        if is_new == 'No':
            for node in [pdp_data, pdp_item, p_enrich]:
                ornaments = node.get("ornaments", [])
                for orn in ornaments:
                    disp = str(orn.get("display", "")).upper()
                    if "NEW" in disp and not any(k in disp for k in ["COLOR", "SIZE", "FABRIC"]):
                        is_new = 'Yes'
                        break
                if is_new == 'Yes':
                    break

        buy_url = p_enrich.get("buy_url")

        # ==================== ✅ 标题清洗：只保留™及以前内容 ====================
        if title:
            tm_index = title.find('™')
            if tm_index != -1:
                title = title[:tm_index+1].strip()
            else:
                title = title.strip()

        # ==================== ✅ 按颜色生成商品（新增尺码汇总） ====================
        color_items = []
        used_colors = set()
        for color in color_list:
            if color in used_colors:
                continue
            used_colors.add(color)
            
            item = {
                "TCIN": tcin,
                "标题": title,
                "品牌": clean_text(p_brand.get("name")),
                "价格": out_current_price,
                "零售价": out_retail,
                "原价": out_reg,
                "促销活动": is_promotion,
                "节省金额": out_save_amt,
                "折扣比例": out_discount_pct,
                "最大折扣": out_max_discount, 
                "清仓状态": is_clearance,
                "材料(汇总)": material_summary,
                "购买次数": sales_count_num,
                "预计送达": delivery_date,
                "新品标签": is_new,
                "评分": rating_avg,
                "评分数量": rating_count,
                "分项评分": sec_ratings,
                "颜色汇总": color_result,
                "颜色": color,
                # ✅ 新增字段
                "尺码汇总": size_result,
                "简洁卖点": " | ".join(clean_bullets),
                "图片链接": p_enrich.get("images", {}).get("primary_image_url"),
                "购买链接": buy_url,
                "商品类型": pdp_item.get("product_classification", {}).get("item_type", {}).get("name")
            }
            item.update(feature_dict)
            color_items.append(item)
        
        reviews = crawl_reviews_enhanced(tcin, title, buy_url)
        return color_items, reviews

    except Exception as e:
        return [], []

# ===================== PART 3: 保存逻辑 =====================

def save_to_csv_for_dbeaver(product_list, review_list, base_filename):
    if not product_list:
        print("⚠️ 没有数据可保存")
        return

    base_name = base_filename.replace('.xlsx', '').replace('.csv', '')
    print(f"💾 正在生成 CSV 文件 (Base: {base_name})...")

    # --- 1. 保存商品主数据 (Products) ---
    df_p = pd.DataFrame(product_list)
    
    fixed_cols = [
        'TCIN', '标题', '品牌', '价格', '零售价', '原价', 
        '促销活动', '节省金额', '折扣比例', '最大折扣', '清仓状态',
        '颜色', '颜色汇总', '尺码汇总', '材料(汇总)', 
        '购买次数',
        '预计送达', '新品标签',
        '评分', '评分数量', '分项评分',
        '简洁卖点', '商品类型', '图片链接', '购买链接'
    ]
    existing_fixed = [c for c in fixed_cols if c in df_p.columns]
    dynamic_cols = [c for c in df_p.columns if c not in fixed_cols]
    final_order = existing_fixed + dynamic_cols
    df_p = df_p[final_order]

    p_filename = f"{base_name}_products.csv"
    df_p.to_csv(p_filename, index=False, encoding='utf-8-sig')
    print(f"✅ 商品数据已保存: {p_filename}")

    # --- 2. 保存评论数据 (Reviews) ---
    if review_list:
        df_r = pd.DataFrame(review_list)
        ideal_order = ["商品TCIN", "商品标题", "星级", "评论带图", "评论标题", "评论内容", "作者", "是否验证购买", "有用数", "提交时间", "商品购买链接"]
        current_cols = df_r.columns.tolist()
        final_r_order = [c for c in ideal_order if c in current_cols] + [c for c in current_cols if c not in ideal_order]
        df_r = df_r[final_r_order]
        
        r_filename = f"{base_name}_reviews.csv"
        df_r.to_csv(r_filename, index=False, encoding='utf-8-sig')
        print(f"✅ 评论数据已保存: {r_filename}")

    # --- 3. 保存统计数据 (Stats) ---
    if not df_p.empty and '商品类型' in df_p.columns:
        def extract_count(val):
            if not val or not isinstance(val, str): return 0
            match = re.search(r'\[共(\d+)种\]', val)
            if match: return int(match.group(1))
            tokens = [t for t in val.split(',') if t.strip()]
            return len(tokens) if tokens else 0

        df_calc = df_p.copy()
        df_calc['商品类型'] = df_calc['商品类型'].fillna('未分类')
        if '颜色汇总' in df_calc.columns:
            df_calc['ColorCount'] = df_calc['颜色汇总'].apply(extract_count)
        else:
            df_calc['ColorCount'] = 0
        
        df_stats = df_calc.groupby('商品类型')[['ColorCount']].mean().reset_index()
        df_stats.columns = ['商品类型', '平均颜色数量']
        df_stats['平均颜色数量'] = df_stats['平均颜色数量'].round(2)
        df_stats = df_stats.sort_values(by='平均颜色数量', ascending=False)

        s_filename = f"{base_name}_stats.csv"
        df_stats.to_csv(s_filename, index=False, encoding='utf-8-sig')
        print(f"✅ 统计数据已保存: {s_filename}")

# ==================== PART 4: 高度定制美化版 Excel 导出 ====================

def save_to_excel_pretty(product_list, review_list, base_filename):
    if not product_list:
        return

    base_name = base_filename.replace('.xlsx', '').replace('.csv', '')
    excel_filename = f"{base_name}.xlsx"
    
    print(f"📊 正在生成美化版 Excel 报表: {excel_filename} ...")
    
    df_p = pd.DataFrame(product_list)
    
    fixed_cols = [
        'TCIN', '标题', '品牌', '价格', '零售价', '原价', 
        '促销活动', '节省金额', '折扣比例', '最大折扣', '清仓状态',
        '颜色', '颜色汇总', '尺码汇总', '材料(汇总)', 
        '购买次数',
        '预计送达', '新品标签',
        '评分', '评分数量', '分项评分',
        '简洁卖点', '商品类型', '图片链接', '购买链接'
    ]
    existing_fixed = [c for c in fixed_cols if c in df_p.columns]
    dynamic_cols = [c for c in df_p.columns if c not in fixed_cols]
    final_order = existing_fixed + dynamic_cols
    df_p = df_p[final_order]

    wb = Workbook()
    
    ws = wb.active
    ws.title = "商品数据"
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    for col_num, column_title in enumerate(df_p.columns, 1):
        cell = ws.cell(row=1, column=col_num, value=column_title)
        cell.font = header_font
        if column_title in fixed_cols:
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        else:
            cell.fill = PatternFill(start_color="538DD5", end_color="538DD5", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=Side(style='thin'))
        
        col_letter = get_column_letter(col_num)
        if '标题' in column_title: ws.column_dimensions[col_letter].width = 50
        elif '卖点' in column_title: ws.column_dimensions[col_letter].width = 40
        elif '评分' in column_title: ws.column_dimensions[col_letter].width = 15
        elif '链接' in column_title: ws.column_dimensions[col_letter].width = 15
        else: ws.column_dimensions[col_letter].width = 15

    for row_idx, row in enumerate(df_p.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            col_name = df_p.columns[col_idx-1]
            
            is_valid_num = isinstance(value, (int, float)) and not math.isnan(value)

            if col_name == '购买次数':
                if is_valid_num and value > 0:
                    cell.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
                    cell.font = Font(bold=True)
                    cell.number_format = '#,##0'
            elif col_name == '预计送达' and value and "无" not in value and "失败" not in value:
                cell.font = Font(color="006100", bold=True)
            elif col_name in ['零售价', '原价', '节省金额']:
                if is_valid_num:
                    cell.number_format = '$#,##0.00'
                    if col_name == '节省金额' and value > 0: cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif isinstance(value, str):
                    cell.alignment = Alignment(horizontal="center")
            elif col_name == '折扣比例':
                if is_valid_num:
                    cell.number_format = '0%'
                    if value > 0: cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                elif isinstance(value, str):
                    cell.alignment = Alignment(horizontal="center")
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            elif col_name == '最大折扣':
                if is_valid_num:
                    cell.number_format = '0%'
                    if value > 0.5: 
                        cell.font = Font(color="FF0000", bold=True)
            elif col_name == '促销活动' and value == 'Yes':
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                cell.font = Font(color="9C0006", bold=True)
            elif col_name == '清仓状态' and value == 'Yes':
                cell.fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
            elif col_name == '评分' and is_valid_num:
                cell.number_format = '0.0'
                if value >= 4.5: cell.font = Font(bold=True, color="006100")
            elif '链接' in col_name and isinstance(value, str) and value.startswith('http'):
                cell.value = "🔗 点击查看"
                cell.hyperlink = value
                font = Font(color="0563C1", underline="single")
                cell.font = font
                cell.alignment = Alignment(horizontal="center")
            
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), bottom=Side(style='thin'))

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

    if review_list:
        ws_r = wb.create_sheet("商品评论")
        df_r = pd.DataFrame(review_list)
        ideal_order = ["商品TCIN", "商品标题", "星级", "评论带图", "评论标题", "评论内容", "作者", "是否验证购买", "有用数", "提交时间", "商品购买链接"]
        current_cols = df_r.columns.tolist()
        final_r_order = [c for c in ideal_order if c in current_cols] + [c for c in current_cols if c not in ideal_order]
        df_r = df_r[final_r_order]

        for col_num, column_title in enumerate(df_r.columns, 1):
            cell = ws_r.cell(row=1, column=col_num, value=column_title)
            cell.font = header_font
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            col_letter = get_column_letter(col_num)
            if '内容' in column_title: ws_r.column_dimensions[col_letter].width = 60
            elif '标题' in column_title: ws_r.column_dimensions[col_letter].width = 30
            elif '链接' in column_title: ws.column_dimensions[col_letter].width = 15
            else: ws_r.column_dimensions[col_letter].width = 15

        for row_idx, row in enumerate(df_r.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws_r.cell(row=row_idx, column=col_idx, value=value)
                col_name = df_r.columns[col_idx-1]
                if '链接' in col_name and isinstance(value, str) and value.startswith('http'):
                    cell.value = "🔗 购买链接"
                    cell.hyperlink = value
                    cell.font = Font(color="0563C1", underline="single")
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), bottom=Side(style='thin'))
        ws_r.auto_filter.ref = ws_r.dimensions

    if not df_p.empty and '商品类型' in df_p.columns:
        ws_s = wb.create_sheet("颜色统计")
        def extract_count(val):
            if not val or not isinstance(val, str): return 0
            match = re.search(r'\[共(\d+)种\]', val)
            if match: return int(match.group(1))
            tokens = [t for t in val.split(',') if t.strip()]
            return len(tokens) if tokens else 0

        df_calc = df_p.copy()
        df_calc['商品类型'] = df_calc['商品类型'].fillna('未分类')
        if '颜色汇总' in df_calc.columns:
            df_calc['ColorCount'] = df_calc['颜色汇总'].apply(extract_count)
        else:
            df_calc['ColorCount'] = 0
        
        df_stats = df_calc.groupby('商品类型')[['ColorCount']].mean().reset_index()
        df_stats.columns = ['商品类型', '平均颜色数量']
        df_stats['平均颜色数量'] = df_stats['平均颜色数量'].round(2)
        df_stats = df_stats.sort_values(by='平均颜色数量', ascending=False)

        headers = ['商品类型', '平均颜色数量']
        for col_num, column_title in enumerate(headers, 1):
            cell = ws_s.cell(row=1, column=col_num, value=column_title)
            cell.font = header_font
            cell.fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws_s.column_dimensions[get_column_letter(col_num)].width = 30 if col_num == 1 else 20

        for row_idx, row in enumerate(df_stats.itertuples(index=False), 2):
            ws_s.cell(row=row_idx, column=1, value=row[0])
            c_cell = ws_s.cell(row=row_idx, column=2, value=row[1])
            c_cell.alignment = Alignment(horizontal="center")
            for col in range(1, 3):
                ws_s.cell(row=row_idx, column=col).border = Border(bottom=Side(style='thin'))

    wb.save(excel_filename)
    print(f"✅ Excel 报表已生成: {excel_filename}")

# ==================== 5. 主流程 ====================

def main():
    print("="*60)
    print("  Target 爬虫 (防封杀版 + 中部时间戳)")
    print("="*60)
    
    input_text = input("🔎 请输入品牌名 (逗号分隔，'all' 爬全部): ").strip()
    if not input_text: return
    target_brands_filter = {b.strip().lower() for b in input_text.split(',')} if input_text != 'all' else None
    
    crawler = TargetCrawler(BASE_API_URL)
    target_cats = crawler.get_target_category_ids()
    if not target_cats: return

    all_tasks = []
    print(f"\n⚡️ 正在扫描品牌列表...")
    for cat in target_cats:
        brands = crawler.get_brands_for_category(cat['id'], cat['name'])
        for b_info in brands:
            if target_brands_filter:
                match = False
                for kw in target_brands_filter:
                    if kw in b_info['brand_name'].lower(): match = True; break
                if not match: continue
            all_tasks.append(b_info)
    
    print(f"📦 筛选出 {len(all_tasks)} 个品牌任务，开始获取商品列表...")
    
    final_products_to_crawl = []
    with ThreadPoolExecutor(max_workers=MAX_WORKERS_LISTING) as executor:
        future_to_task = {executor.submit(crawler.crawl_task_logic, t): t for t in all_tasks}
        for future in as_completed(future_to_task):
            task = future_to_task[future]
            try:
                products = future.result() 
                for p in products:
                    tcin = p.get('tcin')
                    if tcin:
                        final_products_to_crawl.append({
                            'tcin': tcin,
                            'origin_brand': task['brand_name'],
                            'origin_category': task['cat_name'],
                            'plp_data': p 
                        })
            except Exception: pass
            
    unique_map = {p['tcin']: p for p in final_products_to_crawl}
    final_products_to_crawl = list(unique_map.values())
    print(f"\n✅ 累计发现 {len(final_products_to_crawl)} 个唯一商品 TCIN。")
    
    if not final_products_to_crawl:
        print("❌ 未发现商品。")
        return

    print(f"\n🚀 开始深度抓取详情 ({len(final_products_to_crawl)} 个任务)...")
    all_products = []
    all_reviews = []
    
    with ThreadPoolExecutor(max_workers=MAX_WORKERS_DETAIL) as executor:
        future_to_p = {executor.submit(process_single_product, p): p for p in final_products_to_crawl}
        for future in tqdm(as_completed(future_to_p), total=len(final_products_to_crawl), desc="详情解析", unit="个"):
            prods, revs = future.result()
            if prods: 
                all_products.extend(prods)
            if revs:
                all_reviews.extend(revs)

    # 获取 Target 总部 (明尼阿波利斯) 的中部时间，格式化为 "年月日_时分"
    timestamp = pd.Timestamp.now(tz='US/Central').strftime("%Y%m%d_%H%M")
    
    brand_tag = "ALL" if not target_brands_filter else "_".join(list(target_brands_filter))[:20]
    
    # 创建文件夹逻辑 (品牌名_时间戳) 
    output_folder = f"{brand_tag}_{timestamp}" 
    if not os.path.exists(output_folder):
        os.makedirs(output_folder, exist_ok=True)
        print(f"📁 已创建输出目录: {output_folder}")

    # 修改 base_filename，使其包含路径
    file_prefix = f"Target_Data_{brand_tag}_{timestamp}"
    base_filename = os.path.join(output_folder, file_prefix)
    
    # 1. 保存 CSV
    save_to_csv_for_dbeaver(all_products, all_reviews, base_filename)

    # 2. 保存 Excel 
    save_to_excel_pretty(all_products, all_reviews, base_filename)
    
    print("\n🎉 全部完成")

if __name__ == "__main__":
    main()